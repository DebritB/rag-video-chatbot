#!/usr/bin/env python3
"""
Grade video transcripts using offline LLM and create graded Excel file.
Uses Ollama for offline LLM inference.
Criteria:
1. Clarity of explanation of the target problem (Sentiment Analysis)
2. Clarity of demonstration of the process (Algorithm & Evaluation Metrics)
"""

import sys
from pathlib import Path
from typing import Dict, Tuple
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font
import requests
import json
import logging

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

class VideoGrader:
    def __init__(self, input_excel: str, output_excel: str = "graded_videos.xlsx", 
                 model: str = "mistral"):
        self.input_excel = Path(input_excel)
        self.output_excel = Path(output_excel)
        self.ollama_url = "http://localhost:11434/api/generate"
        self.model = model
        
        # Check Ollama connection
        self.check_ollama_connection()
        
        # Color fills
        self.red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        self.yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        self.white_font = Font(bold=True, color="FFFFFF")
        self.black_font = Font(bold=False, color="000000")
    
    def check_ollama_connection(self):
        """Check if Ollama is running."""
        try:
            response = requests.get(self.ollama_url.replace("/api/generate", "/api/tags"), timeout=5)
            if response.status_code == 200:
                logger.info(f"✓ Ollama connected successfully")
                logger.info(f"  Model: {self.model}")
            else:
                logger.error(f"Ollama connection failed (status {response.status_code})")
                logger.error("Make sure Ollama is running: ollama serve")
                sys.exit(1)
        except requests.exceptions.ConnectionError:
            logger.error("✗ Cannot connect to Ollama at " + self.ollama_url)
            logger.error("Start Ollama first:")
            logger.error("  1. Download from https://ollama.ai")
            logger.error("  2. Run: ollama serve")
            logger.error(f"  3. In another terminal: ollama pull {self.model}")
            sys.exit(1)
    
    def query_llm(self, prompt: str) -> str:
        """Query the offline LLM via Ollama."""
        try:
            payload = {
                "model": self.model,
                "prompt": prompt,
                "stream": False,
                "temperature": 0.1  # Low temperature for consistent grading
            }
            
            response = requests.post(self.ollama_url, json=payload, timeout=120)
            if response.status_code == 200:
                result = response.json()
                return result.get("response", "").strip()
            else:
                logger.error(f"LLM query failed: {response.status_code}")
                return "ERROR"
        except Exception as e:
            logger.error(f"LLM error: {e}")
            return "ERROR"
    
    def grade_criteria_1(self, script: str) -> Tuple[str, str]:
        """Grade Criteria 1: Clarity of explanation of the target problem (Sentiment Analysis)."""
        # Take first half of the transcript to check for early algorithm mention
        first_half = script[:len(script)//2]
        
        prompt = f"""Analyze this video transcript (first half) and check if it describes a SENTIMENT ANALYSIS problem with the CORRECT algorithms.

A sentiment analysis problem has these characteristics:
- Mentions sentences, text data, or documents to analyze
- Discusses pre-processing of text/sentences
- Mentions data source, dataset, or data exploration
- Uses ONE of these algorithms: Naive Bayes, Logistic Regression, SVM/LinearSVC, or Decision Tree
- Algorithm MUST be mentioned in the FIRST HALF of the video

IMPORTANT: 
- If the video mentions an algorithm that is NOT one of these 4, it gets NO MARKS.
- If the algorithm is only mentioned in the SECOND HALF, it gets NO MARKS.

Transcript (First Half):
{first_half[:1500]}

Answer these questions:
1. Is there mention of sentences, text, or data for sentiment analysis?
2. Is there discussion of pre-processing or data preparation?
3. Is the data source or dataset mentioned?
4. What algorithms are mentioned in this first half? Are they ONLY from: Naive Bayes, Logistic Regression, SVM/LinearSVC, Decision Tree?

Response format: ONLY output one of: "Full", "No marks"
- Full marks: Clear mention of text/sentence data AND some discussion of pre-processing/data exploration AND algorithms used are ONLY from the 4 listed above AND mentioned early (first half)
- No marks: Either (1) No clear sentiment analysis problem, OR (2) An algorithm is mentioned that is NOT from the 4 listed above, OR (3) Algorithm not mentioned in first half

Output only the mark level:"""
        
        result = self.query_llm(prompt).strip()
        
        # Parse result
        if "Full" in result:
            return "Full", "Sentiment analysis problem with text data and correct algorithms mentioned early"
        else:
            return "No marks", "Not a clear sentiment analysis problem, wrong algorithms, or algorithm not mentioned in first half"
    
    def grade_criteria_2(self, script: str) -> Tuple[str, str]:
        """Grade Criteria 2: Clarity of demonstration of the process (Algorithm & Evaluation Metrics)."""
        prompt = f"""Analyze this video transcript for sentiment analysis algorithms and evaluation metrics.

Transcript:
{script[:1500]}

Check for these specific algorithms and their parameters:

1. NAIVE BAYES - NO parameters needed for full marks
2. LOGISTIC REGRESSION - specific parameters: max_iter, solver
3. SVM / LinearSVC - NO parameters needed for full marks
4. DECISION TREE - specific parameters: criterion, max_depth, min_samples_leaf

Also check for EVALUATION METRICS. Count as mentioned if ANY of these semantic variations appear:
- Accuracy, precision, recall, F1-score, F1, F-measure
- Confusion matrix, ROC curve, AUC, ROC-AUC
- Classification report, performance metrics
- True positive, false positive, true negative, false negative
- TP, FP, TN, FN, TPR, FPR
- Sensitivity, specificity, balanced accuracy
- Cross-validation, validation score, test accuracy
- Loss function, error rate, misclassification
- Model evaluation, performance evaluation, testing

Answer these questions:
1. How many of the above 4 algorithms (Naive Bayes, Logistic Regression, SVM/LinearSVC, Decision Tree) are mentioned?
2. For Logistic Regression (if mentioned): Are parameters like max_iter, solver mentioned?
3. For Decision Tree (if mentioned): Are parameters like criterion, max_depth, min_samples_leaf mentioned?
4. Are ANY evaluation metrics or their semantic variations mentioned?

Based on your analysis, assign a grade:
- Full marks: 1 algorithm mentioned + proper parameters (for LogReg or DecisionTree) or no params needed (NaiveBayes/SVM) + evaluation metrics mentioned (ANY of semantic variations)
- Average: 1 algorithm mentioned + NO parameters mentioned + evaluation metrics mentioned (ANY of semantic variations)
- Fair: 1 algorithm mentioned + NO parameters mentioned (Except Naive Bayes and SVM/LinearSVC, for these no need to mention parameters) OR NO evaluation metrics mentioned
- No marks: Multiple algorithms mentioned OR algorithm is different from the 4 listed above

IMPORTANT: For Naive Bayes and SVM/LinearSVC, parameters are NOT required for any grade level.
For Logistic Regression and Decision Tree, check if parameters are mentioned.

Output format: ONLY the grade (Full/Average/Fair/No marks) on first line, then brief reason on second line"""
        
        result = self.query_llm(prompt).strip()
        lines = result.split('\n')
        grade = lines[0].strip() if lines else "No marks"
        reason = lines[1].strip() if len(lines) > 1 else "Unable to determine"
        
        if "Full" in grade:
            return "Full", reason
        elif "Average" in grade:
            return "Average", reason
        elif "Fair" in grade:
            return "Fair", reason
        else:
            return "No marks", reason
    
    def format_duration_to_seconds(self, duration_str: str) -> float:
        """Convert duration string (MM:SS or H:MM:SS) to seconds."""
        try:
            parts = duration_str.split(':')
            if len(parts) == 2:
                minutes, seconds = map(int, parts)
                return minutes * 60 + seconds
            elif len(parts) == 3:
                hours, minutes, seconds = map(int, parts)
                return hours * 3600 + minutes * 60 + seconds
            return 0.0
        except:
            return 0.0
    
    def grade_transcript(self, video_name: str, script: str, duration: str) -> Dict:
        """Grade a single transcript."""
        logger.info(f"Grading: {video_name}")
        
        # Grade each criteria
        c1_grade, c1_reason = self.grade_criteria_1(script)
        c2_grade, c2_reason = self.grade_criteria_2(script)
        
        # Check video length (> 4:03 = 243 seconds)
        duration_seconds = self.format_duration_to_seconds(duration)
        is_long_video = duration_seconds > 243
        
        logger.info(f"  C1: {c1_grade} | C2: {c2_grade} (Len: {duration})")
        
        return {
            "video_name": video_name,
            "script": script,
            "duration": duration,
            "c1_grade": c1_grade,
            "c1_reason": c1_reason,
            "c2_grade": c2_grade,
            "c2_reason": c2_reason,
            "is_long_video": is_long_video,
        }
    
    def load_transcriptions(self) -> list:
        """Load transcriptions from input Excel file."""
        if not self.input_excel.exists():
            logger.error(f"Input file not found: {self.input_excel}")
            sys.exit(1)
        
        wb = openpyxl.load_workbook(self.input_excel)
        ws = wb.active
        
        transcriptions = []
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            video_name = row[0].value if row[0] else "Unknown"
            script = row[1].value if row[1] else ""
            duration = row[2].value if row[2] else "0:00"
            
            if script and script.strip():
                transcriptions.append({
                    "video_name": video_name,
                    "script": script,
                    "duration": duration
                })
        
        logger.info(f"Loaded {len(transcriptions)} transcriptions from {self.input_excel}")
        return transcriptions
    
    def save_graded_excel(self, results: list):
        """Save graded results to Excel with color coding."""
        logger.info(f"Saving graded results to {self.output_excel}...")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Graded Transcriptions"
        
        # Add headers
        headers = [
            "Video Name",
            "Criteria 1\n(Problem Clarity)",
            "Criteria 2\n(Algorithm & Evaluation)",
            "Duration",
            "Notes"
        ]
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Add data rows
        for row_num, result in enumerate(results, 2):
            ws.cell(row=row_num, column=1).value = result["video_name"]
            ws.cell(row=row_num, column=2).value = result["c1_grade"]
            ws.cell(row=row_num, column=3).value = result["c2_grade"]
            ws.cell(row=row_num, column=4).value = result["duration"]
            
            # Notes
            notes = []
            if result["is_long_video"]:
                notes.append("Video > 4:03 min")
            ws.cell(row=row_num, column=5).value = "; ".join(notes) if notes else ""
            
            # Apply color coding
            # RED if Criteria 1 OR Criteria 2 is "No marks"
            if result["c1_grade"] == "No marks" or result["c2_grade"] == "No marks":
                for col in range(1, 6):
                    cell = ws.cell(row=row_num, column=col)
                    cell.fill = self.red_fill
                    cell.font = self.white_font
            # YELLOW if video > 4:03 min (and not already red)
            elif result["is_long_video"]:
                for col in range(1, 6):
                    cell = ws.cell(row=row_num, column=col)
                    cell.fill = self.yellow_fill
                    cell.font = self.black_font
            
            # Alignment
            for col in range(1, 6):
                ws.cell(row=row_num, column=col).alignment = Alignment(
                    horizontal="left", vertical="top", wrap_text=True
                )
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 25
        
        # Set row heights
        ws.row_dimensions[1].height = 40
        for row_num in range(2, len(results) + 2):
            ws.row_dimensions[row_num].height = None  # Auto-adjust
        
        wb.save(self.output_excel)
        logger.info(f"✓ Graded Excel file saved: {self.output_excel}")
    
    def run(self):
        """Main pipeline: load, grade, and save."""
        logger.info("=" * 60)
        logger.info("VIDEO TRANSCRIPT GRADING PIPELINE (Offline LLM)")
        logger.info("=" * 60)
        
        # Load transcriptions
        transcriptions = self.load_transcriptions()
        
        # Grade each transcript
        results = []
        for idx, trans in enumerate(transcriptions, 1):
            logger.info(f"[{idx}/{len(transcriptions)}] Grading: {trans['video_name']}")
            result = self.grade_transcript(
                trans["video_name"],
                trans["script"],
                trans["duration"]
            )
            results.append(result)
        
        # Save graded results
        self.save_graded_excel(results)
        
        logger.info("=" * 60)
        logger.info("✓ Grading completed successfully!")
        logger.info("=" * 60)


def main():
    input_file = Path(__file__).parent / "transcriptions.xlsx"
    output_file = Path(__file__).parent / "graded_videos.xlsx"
    
    if not input_file.exists():
        logger.error(f"Input file not found: {input_file}")
        logger.error("Run extract_and_transcribe.py first to generate transcriptions.xlsx")
        sys.exit(1)
    
    grader = VideoGrader(str(input_file), str(output_file), model="neural-chat")
    grader.run()


if __name__ == "__main__":
    main()
